"""Generic XLSX export: a posted table comes back as a real .xlsx that round-trips through openpyxl."""
from __future__ import annotations

import io
import types
import unittest

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.main import app
from app.services.auth_service import get_current_user


class ExportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Base.metadata.create_all(self.engine)
        self.Session = sessionmaker(bind=self.engine)
        app.dependency_overrides[get_db] = lambda: (s for s in [self.Session()])
        app.dependency_overrides[get_current_user] = lambda: types.SimpleNamespace(username="t", role="admin")
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()

    def test_xlsx_roundtrip(self):
        r = self.client.post("/export/xlsx", json={
            "sheet_name": "Common", "filename": "report",
            "headers": ["Number", "Count"],
            "rows": [["9876543210", 3], ["111", None]],
        })
        self.assertEqual(r.status_code, 200, r.text)
        self.assertIn("spreadsheetml", r.headers["content-type"])
        self.assertIn("report.xlsx", r.headers.get("content-disposition", ""))
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        self.assertEqual(ws.title, "Common")
        self.assertEqual([c.value for c in ws[1]], ["Number", "Count"])
        self.assertEqual(ws.cell(row=2, column=1).value, "9876543210")
        self.assertEqual(ws.cell(row=2, column=2).value, 3)


if __name__ == "__main__":
    unittest.main()
